
#Program Warning:
print("""Program Warning:
1) Never change the filenames, sheetnames, values, positioning or layout of 'Crypto Database.xlsx' and 'Crypto Valuation.xlsx', as we referenced their rows/columns for data extraction or appendation (to see which rows/columns are 'safe' to be modified, open the respective files)
2) If you obeyed the above but the program suddenly returns an error or doesn't run as you expect, it is likely due to your internet connection, a mismatch in this program's global variable (ex: wrong link, etc.), or that CMC's webpage design has been changed (causing the location of elements and x-paths to change as well). Start debugging from this, as our program's logic has been reviewed multiple times and hence won't be the culprit""")

import time
import openpyxl as opx
import xlwings as xw
import pandas as pd
from tqdm import tqdm
from selenium import webdriver as wd
from selenium.webdriver.chrome.options import Options

#loading the worksheet
directory = '/Users/delvinkennedy/sandbox/Personal Investments/Cryptos/'
template_directory = f'{directory}Crypto Valuation (Template).xlsx'
model_directory = f'{directory}Crypto Valuation.xlsx'
cryptoValuation_template = opx.load_workbook(template_directory)
QTM_Model = cryptoValuation_template['QTM Model']
print('Please wait...')			#just for visual indication that our program has started running

#Retrieving and formatting the data: "Date of Valuation"
def format_datetime_to_excel(secs_since_pythonEpoch):		#Function: Converting python datetime to excel datetime
	excelDay_at_pythonEpoch = 25569		#the 'excel format' for 01/01/1970 (literally means 25569 days passed since 01/01/1900, excel's time reference)	
	pythonDay_since_pythonEpoch = round(secs_since_pythonEpoch / 86400, 0)		#num of secs that has passed since 00:00:00 01/01/1970 (python's time reference), converted to num of days (1 day = 86400 secs)
	excelDay = excelDay_at_pythonEpoch + pythonDay_since_pythonEpoch
	return excelDay
now = time.time()
currentDate_excelFormat = format_datetime_to_excel(now)

#Function: Formatting extracted str data to float
def format_to_float(string):
	formatted_string = string.replace('$','').replace(',','')
	return float(formatted_string)
#Setting up our browser
chromedriver_filepath = '/Users/delvinkennedy/sandbox/Python/chromedriver'
options = wd.ChromeOptions(); options.add_argument("--headless")
browser = wd.Chrome(chromedriver_filepath, options=options)

#Loads data from "Crypto Database.xlsx", sums the count of cryptos in "Buylist" or "Shortlisted" (to determine the number of cryptos included in our valuation model), and extracts the "Network Name" of each included crypto (so that we can get its CMC link and extract their market cap and price)
startingRowIndex = 23; startingColIndex = 3 		#indexing starts from 0 (col 3 means D, col 5 means F, row 5 means row 6, etc.)
df_cryptodatabase = pd.read_excel(f"{directory}Crypto Database.xlsx", skiprows=range(0, startingRowIndex-1), usecols=range(startingColIndex, startingColIndex+1))		#extracts the "Status" column from our crypto database and returns it in form of dataframe
df_allStatus = df_cryptodatabase.value_counts()			#returns a series as a result of the "value_counts()" operation on the extracted dataframe
buylist_count = df_allStatus.get(key='Buylist')
buylist_count = 0 if (buylist_count == None) else buylist_count
shortlist_count = df_allStatus.get(key='Shortlisted')
shortlist_count = 0 if (shortlist_count == None) else shortlist_count
valuedCryptos_count = buylist_count + shortlist_count			#sums the counts of "Buylist" and "Shortlisted"
cryptos = []
for i in range(valuedCryptos_count):
	crypto = QTM_Model.cell(column=2, row=34+i)
	cryptos.append(crypto.value)		#extracts "Network Name" of included cryptos and appends them to a list
#Retrieving and formatting the data: "Market Cap (USD)" and "Price (USD)" of each crypto on our valuation model
price_list = []
crypto_baselink = 'https://coinmarketcap.com/currencies/'
for crypto in tqdm(cryptos, desc='Progress'):
	crypto_formatted = crypto.lower().replace(' ', '-')			#based on CMC's pattern for each crypto webpage (ex: https://coinmarketcap.com/currencies/near-protocol for NEAR protocol, etc.)
	crypto_link = f'{crypto_baselink}{crypto_formatted}'
	browser.implicitly_wait(5); browser.get(crypto_link)
	price_raw = browser.find_element_by_css_selector('div.priceValue').text
	price = format_to_float(price_raw)
	price_list.append(price)

#Quits the browser, appends the data to "Crypto Valuation.xlsx", and saves the worksheet
browser.quit()
QTM_Model['D19'].value = currentDate_excelFormat
for rowIndex, price in enumerate(price_list):
	QTM_Model.cell(column=3, row=34+rowIndex).value = price
cryptoValuation_template.save(model_directory)
print('"Crypto Valuation.xlsx" has been created.')
