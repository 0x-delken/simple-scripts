
#Program Warning:
print("""Program Warning:
1) Never change the filenames, sheetnames, values, positioning or layout of 'Stock Database.xlsx', 'Stock Valuation.xlsx', and 'Stock Valuation (Root).xlsx', as we referenced their rows/columns for data extraction or appendation
2) Never change the directories inside "Personal Investments/Stocks/", as we use the os module to create or delete files for our shortlisted stocks
3) If you obeyed the above but the program suddenly returns an error or doesn't run as you expect, it is likely due to your internet connection, a mismatch in this program's global variable (ex: wrong link, change in WSJ's financial statement entry names, etc.), or that investing.com's webpage design has been changed (ex: new pop-ups, change in element locations, etc.). Start debugging from this, as our program's logic has been reviewed multiple times and hence won't be the culprit
NB: When "N run" is done, there will always be 'pending' stocks (denoted in "Additional Notes" column in 'Stock Database.xlsx'). Check them manually in our stock database to diagnose the problem (is the ticker wrong, does WSJ provide data for that stock, or is it just WSJ's server or our internet?). Once done, "C run" this program to loop over and 'process' them""")

print('Program is running. Please wait...')
import os, time, re
import datetime as dt
import send2trash as s2t
import openpyxl as opx
import xlwings as xw
import numpy as np
import pandas as pd
import warnings; warnings.filterwarnings('ignore', message="rubicon.objc.ctypes_patch has only been tested with Python 3.4 through 3.8. You are using Python 3.9.5. Most likely things will work properly, but you may experience crashes if Python's internals have changed significantly.")
import pyautogui as pyag
import pyinputplus as pyip
from tqdm import tqdm
from selenium import webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support import expected_conditions as EC

#Custom exception class to handle intentionally thrown exceptions
#NB: By raising an exception, code execution will stop there and the function will be 'killed'. Custom exceptions are so that we can append data to stock database based on the exception that is raised (ex: SkipValuationException is to append "Skipped" on 'Status' column)
class CannotFetchDataException(Exception):
	pass
class SkipValuationException(Exception):
	pass

#Function: The Main Valuation Program
def stockValuation(stock_name, stock_ticker, country_code, valuation_type='New Valuation', existing_status=None):
	#Launching the browser, setting up financial statement links, and declare lists to hold financial data (fiscal year, statements, etc.)
	browser = wd.Chrome(chromedriver_filepath, desired_capabilities=capa)			#include "desired_capabilities=capa" to enable 'force stop' functionality
	statements = ['Income Statement', 'Balance Sheet', 'Cash Flow Statement']
	incomeStatement_link = f'https://www.wsj.com/market-data/quotes/{country_code}/{stock_ticker}/financials/annual/income-statement'
	balanceSheet_link = f'https://www.wsj.com/market-data/quotes/{country_code}/{stock_ticker}/financials/annual/balance-sheet'
	cashflowStatement_link = f'https://www.wsj.com/market-data/quotes/{country_code}/{stock_ticker}/financials/annual/cash-flow'
	financialStatements = {statements[0]: incomeStatement_link, statements[1]: balanceSheet_link, statements[2]: cashflowStatement_link}
	stock_price = []; stock_fiscalYear = []; stock_financials = []
	#Function to force stop WSJ page loading (if not handled, our code will get stuck in "browser.get()" until WSJ's webpage completely loads, which takes forever)
	#NB: The 4 lines of code below is 'default syntax' (you can't put it inside try-except block, or anything else; you can only write these 4 lines row-by-row as-is)
	def forceStop_handler():
		testElement_xpath = '//*[@id="navCol"]/div[2]'
		wait = WebDriverWait(browser, 5)
		browser.get(link)
		try:
			wait.until(EC.presence_of_element_located((By.XPATH, testElement_xpath)))
		except TimeoutException:
			pass
		else:
			time.sleep(5)		#to allow the element up to 5 secs to load completely if it is present (remember, present element doesn't mean loaded element)
		browser.execute_script("window.stop();")		#the 'browser force stop' command
	#Function to handle WSJ's "data not available" instances
	def pullData_from_WSJ(element_xpath):
		for attempts in range(8):			#we allow up to 8 attempts to retrieve the financial statements
			try:
				data = browser.find_element_by_xpath(element_xpath).text
			except (NoSuchElementException, StaleElementReferenceException):		#if element not found, tests for whether it is a "404 Page Not Found" error
				try:
					browser.find_element_by_xpath('//*[@id="main"]/div/div/div/div[1]')			#the element that pops up if WSJ returns a "404 Page Not Found" (is caused when the country code or stock ticker is incorrect)
				except (NoSuchElementException, StaleElementReferenceException):
					forceStop_handler()
				else:
					browser.quit()
					raise CannotFetchDataException 				#if "try" block succeeds, it means the page is "404 Not Found" and we directly raise "CannotFetchDataException"
			else:
				return data
				break
		else:
			browser.quit()
			raise CannotFetchDataException
	#Function to convert string to float on extracted financials; ex: (12,840) to -12840, 65,432.1 to 65432.1, etc.
	def convert_str_to_float(element):
		string_element = str(element)
		formatted = string_element.replace(',', '').replace('-', 'nan').translate(string_element.maketrans('(', '-', ')'))			#multiple replace is based on the order of priority (we need to replace '-' with 'nan' first before translating '()' to '-', as we don't want -21 to be 'nan21')
		return float(formatted)

	#Loop to separate financial statements
	for statement in financialStatements:
		#Extracting full data based on statement type (WSJ has different layouts for each statement type)
		link = financialStatements[statement]
		forceStop_handler()
		if statement == statements[0]:
			allData = pullData_from_WSJ('//*[@id="cr_cashflow"]/div[2]/div/table')
			allData_list = allData.split('\n')			#splits the raw text data into newlines
			miscData = allData_list[0]			#the data to be extracted separately (as it only contains the stock's fiscal year, financial data denomination info, and "years_list")
			financialData = allData_list[1:]			#the data to be put into our DataFrame
		elif statement == statements[1]:
			financialData = []
			miscData_list = []
			for attempts in range(3):			#we need to error handle 'expanded data' from balance sheet and cash flow statements separately, as they can't be handled with "pullData_from_WSJ" (only for 'in-place' data)
				allData1 = pullData_from_WSJ('//*[@id="cr_cashflow"]/div[2]/div[2]/table')
				expand_button = browser.find_element_by_xpath('//*[@id="cr_cashflow"]/div[3]/div[1]/h2')
				browser.execute_script("arguments[0].click();", expand_button)
				time.sleep(0.5)			#to give time for our 'expanded' data to load properly
				try:
					allData2 = browser.find_element_by_xpath('//*[@id="cr_cashflow"]/div[3]/div[2]/table').text			#"Liabilities and Shareholders' Equity" expanded data
				except:					#so that if for some reason the expand button doesn't work (and the element isn't extracted properly), we let the loop run (which executes "pullData_from_WSJ" again until "attempts" are exhausted)
					pass
				else:
					allData1_list = allData1.split('\n'); allData2_list = allData2.split('\n')
					miscData_list.append(allData1_list[0])
					financialData1 = allData1_list[1:]; financialData2 = allData2_list[1:]
					financialData.extend(financialData1); financialData.extend(financialData2)
					if len(financialData2) != 0:
						break			#this is to act as a 'safeguard' if WSJ's financial statement table element is detected, but for some reason Selenium can't extract its text (can happen since we didn't extract it using "pullData_from_WSJ"); hence we only break the loop if Selenium extracts its text successfully (proved by this command)
			else:
				browser.quit()
				raise CannotFetchDataException
			miscData = miscData_list[0]
		elif statement == statements[2]:
			financialData = []
			miscData_list = []
			for attempts in range(3):
				allData1 = pullData_from_WSJ('//*[@id="cr_cashflow"]/div[2]/div[2]/table')
				expand_button1 = browser.find_element_by_xpath('//*[@id="cr_cashflow"]/div[3]/div[1]/h2')
				browser.execute_script("arguments[0].click();", expand_button1)
				time.sleep(0.5)
				try:
					allData2 = browser.find_element_by_xpath('//*[@id="cr_cashflow"]/div[3]/div[2]/table').text			#"Cash Flow from Investing Activities" expanded data
					expand_button2 = browser.find_element_by_xpath('//*[@id="cr_cashflow"]/div[4]/div[1]/h2')
					browser.execute_script("arguments[0].click();", expand_button2)
					time.sleep(0.5)
					allData3 = browser.find_element_by_xpath('//*[@id="cr_cashflow"]/div[4]/div[2]/table').text			#"Cash Flow from Financing Activities" expanded data
				except:
					pass
				else:
					allData1_list = allData1.split('\n'); allData2_list = allData2.split('\n'); allData3_list = allData3.split('\n')
					miscData_list.append(allData1_list[0])
					financialData1 = allData1_list[1:]; financialData2 = allData2_list[1:]; financialData3 = allData3_list[1:]
					financialData.extend(financialData1); financialData.extend(financialData2); financialData.extend(financialData3)
					if (len(financialData2) != 0) and (len(financialData3) != 0):
						break
			else:
				browser.quit()
				raise CannotFetchDataException
			miscData = miscData_list[0]
		#Extracting stock price, fiscal year, denomination of financial data, and "years_list"
		if statement == statements[0]:
			stockPrice_raw = browser.find_element_by_xpath('//*[@id="quote_val"]').text
			stockPrice = float(stockPrice_raw.replace(',', ''))
			if country_code == 'UK':
				stockPrice = stockPrice / 100 			#specially for UK, as their stock prices are denominated in pence instead of GBP
			stock_price.append(stockPrice)
			fiscalYear_regex = re.compile(r"\w+[-]\w+")
			fiscalYear_result = fiscalYear_regex.findall(miscData)
			fiscalYear = fiscalYear_result[0]
			stock_fiscalYear.append(fiscalYear)
		denomination_regex = re.compile(r"(Millions|Thousands)")		#searches for "Millions" or "Thousands" within the extracted text
		denomination_result = denomination_regex.findall(miscData)
		denomination = denomination_result[0]
		years_regex = re.compile(r"\d{4}")
		years_list = years_regex.findall(miscData)
		#Setting up objects to extract financials
		labels_regex = re.compile(r"[a-zA-Z&/\-]+")
		values_regex = re.compile(r"""									#start of regex
									(\()?								#start of negative accounting values
									(\d{3})?(\d{2})?(\d{1})?			#millions
									([,])?								#comma separator
									(\d{3})?(\d{2})?(\d{1})?			#thousands
									([,])?								#comma separator
									(\d{3})?(\d{2})?(\d{1})?([\-])?		#one-digit to hundreds, or blank
									([.])?								#dot separator
									(\d{3})?(\d{2})?(\d{1})?			#decimal places
									(\))?								#end of negative accounting values
									""", re.VERBOSE)					#end of regex (with end syntax: re.VERBOSE)
		labels = []; labels_values = []
		#Extracting financials
		for dataByRow in financialData:
			dataByRow_list = dataByRow.split()			#splits by ' ', and stored in form of a list
			#Extracts financials description (ex: "Sales/Revenue", "EBIT", etc.)
			label = []
			for data in dataByRow_list:	
				words = labels_regex.findall(data)			#matches for regex on each 'splitted data'
				try:
					word = words[0]
				except IndexError:			#if regex doesn't match, the list "words" won't form and hence if we call "words[0]" on it, will return an IndexError
					pass
				else:
					label.append(word)			#if regex matches, then append "word" into the list "label"
			label_formatted = ' '.join(label).rstrip(' -')			#removes the ' -' from the right side of our extracted label, as our regex will also take in 'individual' "-" that is meant for blank values; this is why we exclude "(", ")", ".", and ",", as these chars are used predominantly in values
			#Extracts financials values
			raw_data = values_regex.findall(dataByRow)
			values = []
			for i in range(len(label_formatted), len(raw_data)):			#we start iterating for values from the str location after the label (ex: if "Sales/Revenue", we start iterating from 13)
				data = ''.join(raw_data[i])
				values.append(data)
			values_formatted = []
			for value in values:
				if value != '':
					values_formatted.append(value)			#joins together all values that are matched by our regex (including blank "-" values) for each column
			label_value = np.array([values_formatted])
			if len(values_formatted) == 5:			#appends only if the length of "values_formatted" is equal to the number of columns (to discard those numbers that we don't want to include on our dataframe, but are 'accidentally' retrieved by our regex; ex: discarding -21.12 retrieved from -21.12%, etc.)
				labels.append(label_formatted)
				labels_values.append(label_value)
		#Organizing extracted financials into a dataframe, converting it from str to float, formats the row and col indexes of the dataframe, multiply values based on denomination, and appends the dataframe to the 'placeholder list' "stock_financials"
		if len(labels_values) > 0:
			financials = np.vstack(labels_values)
		else:			#to skip valuation on stocks in which its data isn't provided by WSJ (in this case, WSJ only returns financial statement table with no data)
			browser.quit()
			raise SkipValuationException
		df_financials = pd.DataFrame(financials)
		num_of_cols = df_financials.shape[1]
		financials_by_year = []
		for col in range(num_of_cols):			#convert from str to float by column (ex: 2020 column, 2019 column, etc.)
			series_financials = df_financials.iloc[0:, col]
			series_financials = series_financials.apply(convert_str_to_float)
			financials_by_year.append(series_financials)
		df_financials_formatted = pd.DataFrame([])
		for i, year in enumerate(years_list):
			df_financials_formatted[year] = financials_by_year[i]			#appends "financials_by_year" list to our dataframe by column
		df_financials_formatted[statement] = labels			#appends labels to our "statement" column (ex: 'Income Statement', etc.)
		df_financials_formatted = df_financials_formatted.set_index(statement)			#sets our "statement" column as the dataframe's row index
		if denomination == 'Millions':
			df_financials_formatted = df_financials_formatted * 1000			#multiplies all numbers in the dataframe by 1000 if denomination is in 'Millions' (the smallest denomination in WSJ is 'Thousands', hence this is to make sure that all data pulled from WSJ will be represented in terms of 'Thousands')
		stock_financials.append(df_financials_formatted)
	#Quits browser, and assigns dataframes to its respective variables
	browser.quit()
	incomeStatement = stock_financials[0]
	balanceSheet = stock_financials[1]
	cashflowStatement = stock_financials[2]

	#Indexing from the full dataframe to a new dataframe so that we only include select entries and leave out the rest
	#NB: We assume WSJ's financial statement entry names to be exactly similar with the names on the three lists below, and WSJ's financial statement ordering to be from MRY left-to-right (though this only affects "fix_dilutedShares" function)
	#NB: Any change to the POSITIONS on the three lists below must be reflected on "fixing required fields" (just find "_entry" to search for all the variables that calls data directly from these three lists)
	incomeStatement_entries = ['Sales/Revenue', 'Cost of Goods Sold (COGS) incl. D&A', 'SG&A Expense',
							'Equity in Affiliates', 'Minority Interest Expense', 'Income Tax', 'Gross Interest Expense',
							'Net Income Available to Common', 'Diluted Shares Outstanding',
							'Depreciation', 'Depreciation & Amortization Expense']
	balanceSheet_entries = ['Cash & Short Term Investments', 'Total Accounts Receivable', 'Inventories', 'Net Property, Plant & Equipment',
							'Accounts Payable', 'ST Debt & Current Portion LT Debt', 'Long-Term Debt',
							'Total Assets', 'Total Liabilities', "Total Shareholders' Equity",
							'Total Equity', 'Accumulated Minority Interest']
	cashflowStatement_entries = ['Depreciation and Depletion', 'Changes in Working Capital', 'Net Operating Cash Flow',
							'Capital Expenditures', 'Net Investing Cash Flow', 'Cash Dividends Paid - Total',
							'Depreciation, Depletion & Amortization']
	#Leave out "(", ")", ".", and "," from the exact WSJ fields (ex: "incl. (COGS)" becomes "incl COGS"), as our label regex doesn't extract those chars into our dataframe (which makes up our row index)
	incomeStatement_regexEntries = [' '.join(labels_regex.findall(entry)) for entry in incomeStatement_entries]
	balanceSheet_regexEntries = [' '.join(labels_regex.findall(entry)) for entry in balanceSheet_entries]
	cashflowStatement_regexEntries = [' '.join(labels_regex.findall(entry)) for entry in cashflowStatement_entries]
	#Specifying required entries: If one of these entries are not provided, we skip valuation
	required_entries = (incomeStatement_regexEntries[0], incomeStatement_regexEntries[1], incomeStatement_regexEntries[2], incomeStatement_regexEntries[-4], incomeStatement_regexEntries[-3], balanceSheet_regexEntries[0], balanceSheet_regexEntries[3], balanceSheet_regexEntries[-5], balanceSheet_regexEntries[-4], balanceSheet_regexEntries[-2], cashflowStatement_regexEntries[3])
	#Setting up 'helper iterator' to loop between financial statements
	statementIterator = [(incomeStatement, incomeStatement_regexEntries), (balanceSheet, balanceSheet_regexEntries), (cashflowStatement, cashflowStatement_regexEntries)]
	formattedStatements = []			#list placeholder so that we can retain data outside of loop
	for item in statementIterator:
		selectedStatement = item[0]
		selectedEntries = item[1]
		num_of_rows = len(selectedEntries)
		num_of_cols = len(selectedStatement.columns)
		formattedStatement = pd.DataFrame(np.full((1, num_of_cols), np.nan))		#dataframe placeholder so that we can append values to it by using ".loc"
		formattedStatement.columns = selectedStatement.columns
		for row in range(num_of_rows):
			entryName = selectedEntries[row]
			try:
				entryValues = selectedStatement.loc[entryName]
			except KeyError:
				if entryName in required_entries:		#if KeyError, checks whether "entryName" is in 'required entries' (Revenue, COGS, PP&E, etc.); if yes, the stock is probably an FIG or a Real Estate firm (or worse, the data is simply not provided), hence we skip valuation
					raise SkipValuationException
				else:
					entryValues = np.array([np.nan] * num_of_cols).T 			#if the "entryName" is not Revenue, COGS, or SG&A, leave the fields blank (probably the KeyError from "entryName" is on something like 'Equity in Affiliates' or 'Minority Interest Expense')
			formattedStatement.loc[entryName] = entryValues
		formattedStatement = formattedStatement.drop(index=[0])			#to drop the very first row of the dataframe (created as a byproduct of our dataframe placeholder)
		formattedStatements.append(formattedStatement)
	#Assigning extracted data to its respective variables
	stock_price = stock_price[0]
	stock_fiscalYear = stock_fiscalYear[0]
	incomeStatement_formatted = formattedStatements[0]
	balanceSheet_formatted = formattedStatements[1]
	cashflowStatement_formatted = formattedStatements[2]

	#Getting the save directory before applying 'fixes' (so that we can delete 'revaluation stocks' if they raised a SkipValuationException)
	country = []
	for index, element in enumerate(countryType_table['Country']):
		if element == country_code:
			countryType = countryType_table.loc[index, 'Type']
			country.append(country_code)
			break
	else:
		countryType = 'Emerging'
		country = ['EU']
	stock_directory = f'{directory}/Shortlisted Stocks/{countryType}/{country[0]}/{stock_name}'
	#Function to delete 'revaluation stocks' if they raised a SkipValuationException during 'fixing'
	def delete_directory_if_revaluation_stock(directory):
		if valuation_type == 'Revaluation':
			try:
				s2t.send2trash(directory)
			except:
				pass

	#'Fixing' required fields that may have empty values, or that must undergo a certain operation
	#NB: Any dataframes or series that are created inside a function are kept local (hence why pandas has "apply()" to 'accomodate' functions); if your function won't be used in "apply()", then you must return a list before converting it into a dataframe 'outside' the function
	#Income Statement
	revenue_entry = incomeStatement_regexEntries[0]
	COGS_entry = incomeStatement_regexEntries[1]
	SGA_entry = incomeStatement_regexEntries[2]
	dilShares_entry = incomeStatement_regexEntries[-3]
	depreciationIS_entry = incomeStatement_regexEntries[-2]
	depreciationAmortizationIS_entry = incomeStatement_regexEntries[-1]
	#Balance Sheet
	PPE_entry = balanceSheet_regexEntries[3]
	shareholdersEquity_entry = balanceSheet_regexEntries[-3]
	totalEquity_entry = balanceSheet_regexEntries[-2]
	accumulatedMinorityInterest_entry = balanceSheet_regexEntries[-1]
	#Cash Flow Statement
	depreciation_entry = cashflowStatement_regexEntries[0]
	depreciationAmortizationCFS_entry = cashflowStatement_regexEntries[-1]
	#Income Statement - COGS and SG&A not provided: Allocate based on past ratios. Current Year Diluted Shares not provided: Use prev year's diluted shares
	def fix_with_revenue_as_anchor(selected_series):
		revenue_series = incomeStatement_formatted.loc[revenue_entry]
		newValues = []; fixNotes = []
		if selected_series.isna().sum() <= 1:		#we skip valuation if COGS or SG&A isn't provided for two years
			for index, value in enumerate(selected_series):
				if np.isnan(value) == False:
					newValue = value
				elif np.isnan(value) == True:
					subjectValue = revenue_series[index]
					num_of_cols = len(selected_series)
					cols_index = [x for x in range(num_of_cols) if x != index]
					numerator = []; divisor = []
					for i in cols_index:
						numerator.append(selected_series[i])
						divisor.append(revenue_series[i])
					percentageOfRevenue_series = pd.Series(numerator) / pd.Series(divisor)
					newValue = round(percentageOfRevenue_series.mean() * subjectValue, 2)
					years_list = selected_series.index
					notes = f'{selected_series.name} in {years_list[index]} is "fixed" by its Average % to {revenue_series.name}'
					fixNotes.append(notes)
				newValues.append(newValue)
		else:
			delete_directory_if_revaluation_stock(stock_directory)
			raise SkipValuationException
		returnedData = (newValues, fixNotes)
		return returnedData
	def fix_dilutedShares(dilShares_series):
		values = np.array(dilShares_series)
		newValues = []; fixNotes = []
		for i in range(2):			#we skip valuation if there are no diluted shares outstanding figure for two years
			if np.isnan(values[i]) == False:
				newValue = values[i]
				filler_values = np.array([np.nan] * (len(dilShares_series) - 1))
				if i == 0:
					newValues.append(newValue)
					newValues.extend(filler_values)
					break
				elif i == 1:
					newValues.append(newValue)
					newValues.extend(filler_values)
					years_list = dilShares_series.index
					notes = f"{years_list[0]} {dilShares_series.name} figure is brought forward from its previous year's figure ({years_list[1]})"
					fixNotes.append(notes)
					break
		else:
			delete_directory_if_revaluation_stock(stock_directory)
			raise SkipValuationException
		returnedData = (newValues, fixNotes)
		return returnedData
	COGS_returnedData = fix_with_revenue_as_anchor(incomeStatement_formatted.loc[COGS_entry])
	COGS_newValues = COGS_returnedData[0]; COGS_fixNotes = COGS_returnedData[1]
	SGA_returnedData = fix_with_revenue_as_anchor(incomeStatement_formatted.loc[SGA_entry])
	SGA_newValues = SGA_returnedData[0]; SGA_fixNotes = SGA_returnedData[1]
	dilShares_returnedData = fix_dilutedShares(incomeStatement_formatted.loc[dilShares_entry])
	dilShares_newValues = dilShares_returnedData[0]; dilShares_fixNotes = dilShares_returnedData[1]
	#Balance Sheet - Subtract Total Equity with Accumulated Minority Interest
	def fix_shareholdersEquity(totalEquity_series, accumulatedMinorityInterest_series):
		accumulatedMinorityInterest_formattedSeries = accumulatedMinorityInterest_series.replace(np.nan, 0)			#we can directly reference "np.nan" because the series is already in float format (if not, then need to replace with "nan" and convert it to float 'manually')
		shareholdersEquity_series = totalEquity_series - accumulatedMinorityInterest_formattedSeries
		shareholdersEquity = np.array(shareholdersEquity_series)
		return list(shareholdersEquity)
	shareholdersEquity_newValues = fix_shareholdersEquity(balanceSheet_formatted.loc[totalEquity_entry], balanceSheet_formatted.loc[accumulatedMinorityInterest_entry])
	#Cash Flow Statement - Depreciation and Depletion not provided: Use "Depreciation" from IS. If also not provided, use "Depreciation, Depletion & Amortization" from CFS. If again not provided, use "Depreciation & Amortization Expense" from IS. If still not provided, allocate based on past ratios (PPE as the 'anchor', like how Revenue is the 'anchor' for COGS and SG&A)
	def fix_depreciation(depreciation_series):
		depreciationIS_series = incomeStatement_formatted.loc[depreciationIS_entry]
		depreciationAmortizationCFS_series = cashflowStatement_formatted.loc[depreciationAmortizationCFS_entry]
		depreciationAmortizationIS_series = incomeStatement_formatted.loc[depreciationAmortizationIS_entry]
		alternative_series = [depreciationIS_series, depreciationAmortizationCFS_series, depreciationAmortizationIS_series]
		PPE_series = balanceSheet_formatted.loc[PPE_entry]
		newValues = []; fixNotes = []
		newValues_placeholder = []
		for index, value in enumerate(depreciation_series):
			if np.isnan(value) == False:
				newValues_placeholder.append(value)
			elif np.isnan(value) == True:
				for series in alternative_series:
					value = series[index]
					if np.isnan(value) == False:
						newValues_placeholder.append(value)
						years_list = depreciation_series.index
						notes = f"{depreciation_series.name} in {years_list[index]} is obtained from {series.name}"
						fixNotes.append(notes)
						break
				else:
					value = 'No Alternative'
					newValues_placeholder.append(value)
		newValues_filter = (pd.Series(newValues_placeholder) == 'No Alternative')
		if newValues_filter.sum() <= 1:		#we skip valuation if we can't find an alternative for the blank Depreciation value, and if it is also not provided for two years
			for index, value in enumerate(newValues_placeholder):
				if value != 'No Alternative':
					newValue = value
				elif value == 'No Alternative':
					subjectValue = PPE_series[index]
					num_of_cols = len(newValues_placeholder)
					cols_index = [x for x in range(num_of_cols) if x != index]
					numerator = []; divisor = []
					for i in cols_index:
						numerator.append(newValues_placeholder[i])
						divisor.append(PPE_series[i])
					newValue = round((pd.Series(numerator).sum() / pd.Series(divisor).sum()) * subjectValue, 2)
					years_list = depreciation_series.index
					notes = f'{depreciation_series.name} in {years_list[index]} is "fixed" by its Cumulative % to {PPE_series.name}'
					fixNotes.append(notes)
				newValues.append(newValue)
		else:
			delete_directory_if_revaluation_stock(stock_directory)
			raise SkipValuationException
		returnedData = (newValues, fixNotes)
		return returnedData
	depreciation_returnedData = fix_depreciation(cashflowStatement_formatted.loc[depreciation_entry])
	depreciation_newValues = depreciation_returnedData[0]; depreciation_fixNotes = depreciation_returnedData[1]
	#Applying the 'fixes' to our formatted dataframes and compiling all "fixNotes"
	incomeStatement_formatted.loc[COGS_entry] = COGS_newValues
	incomeStatement_formatted.loc[SGA_entry] = SGA_newValues
	incomeStatement_formatted.loc[dilShares_entry] = dilShares_newValues
	balanceSheet_formatted.loc[shareholdersEquity_entry] = shareholdersEquity_newValues
	cashflowStatement_formatted.loc[depreciation_entry] = depreciation_newValues
	fixNotes = COGS_fixNotes + SGA_fixNotes + dilShares_fixNotes + depreciation_fixNotes
	fixNotes_formatted = pd.DataFrame(fixNotes)
	#Cleaning up our formatted dataframes (replacing "np.nan" with 0, and removing non-needed rows) so that our data is ready to be appended to our excel model"
	#NB: Non-needed rows are the 'bottom list' on all of our statement entries (see the user-inputted lists of "incomeStatement_entries", "balanceSheet_entries", and "cashflowStatement_entries")
	incomeStatement_formatted = incomeStatement_formatted.replace(np.nan, 0)
	balanceSheet_formatted = balanceSheet_formatted.replace(np.nan, 0)
	cashflowStatement_formatted = cashflowStatement_formatted.replace(np.nan, 0)
	incomeStatement_formatted = incomeStatement_formatted.drop(index=[depreciationIS_entry, depreciationAmortizationIS_entry])
	balanceSheet_formatted = balanceSheet_formatted.drop(index=[totalEquity_entry, accumulatedMinorityInterest_entry])
	cashflowStatement_formatted = cashflowStatement_formatted.drop(index=[depreciationAmortizationCFS_entry])

	#Appending data to excel valuation model (single values)
	book = opx.load_workbook(valuationTemplate_directory)
	financials_sheet = book['Financials']
	valuation_sheet = book['Valuation']
	financials_sheet['J3'].value = stock_ticker
	financials_sheet['J4'].value = country_code
	valuation_sheet['F4'].value = stock_price
	#Appending data to excel valuation model (dataframe)
	writer = pd.ExcelWriter(valuationModel_rootAccess_directory, engine='openpyxl')
	writer.book = book; writer.sheets = {ws.title: ws for ws in book.worksheets}
	incomeStatement_startRow = 2; balanceSheet_startRow = 13; cashflowStatement_startRow = 25; startCol = 2
	fixNotes_startRow = 36; fixNotes_startCol = 1
	#Raw statements to each respective sheets (for reference only)
	for sheetname in writer.sheets:
		incomeStatement.to_excel(writer, sheet_name='Income Statement', startrow=0, startcol=0, index=True, header=True)
	for sheetname in writer.sheets:
		balanceSheet.to_excel(writer, sheet_name='Balance Sheet', startrow=0, startcol=0, index=True, header=True)
	for sheetname in writer.sheets:
		cashflowStatement.to_excel(writer, sheet_name='Cash Flow Statement', startrow=0, startcol=0, index=True, header=True)
	#Formatted statements and 'fixNotes' to "financials" sheet
	for sheetname in writer.sheets:
		incomeStatement_formatted.to_excel(writer, sheet_name='Financials', startrow=incomeStatement_startRow, startcol=startCol, index=False, header=True)
	for sheetname in writer.sheets:
		balanceSheet_formatted.to_excel(writer, sheet_name='Financials', startrow=balanceSheet_startRow, startcol=startCol, index=False, header=True)
	for sheetname in writer.sheets:
		cashflowStatement_formatted.to_excel(writer, sheet_name='Financials', startrow=cashflowStatement_startRow, startcol=startCol, index=False, header=True)
	for sheetname in writer.sheets:
		fixNotes_formatted.to_excel(writer, sheet_name='Financials', startrow=fixNotes_startRow, startcol=fixNotes_startCol, index=False, header=False)
	#Saving the file with "Stock Valuation Model (Root)" as its name (and closing all associated variables to save resources)
	writer.save()
	writer = None; book = None 				#assign "None" to the variable to 'manually' clear up memory that is holding the variable

	#'Manually' open-save-close our recently saved stock valuation model through xlwings (so that excel will run the formulas inside the workbook and we can extract its resulting value)
	#NB: xlwings is like Selenium for excel (simulates a real user), while openpyxl is like bs4 for excel (purely machine language). Hence, openpyxl is significantly faster than xlwings, although xlwings wins out on functionality
	#NB: Due to xlwings needing to open the excel app directly to modify xlsx, it can only be run through terminal (on first run, click on allow when prompted on permissions), since ST3 doesn't have the necessary permissions to open and close apps
	def open_excel_by_xlwings(directory):			#default syntax to 'manually' open-save-close file; at first run you'll be prompted by excel to "Grant Access" (hence why its important for us to have a 'placeholder excel file' with the same name on the same directory so that we won't be prompted again)
		app = xw.App(visible=False)
		book = app.books.open(directory)
		book.save(directory)
		book.close(); book = None
		app.quit()
	open_excel_by_xlwings(valuationModel_rootAccess_directory)
	#Re-opens the stock valuation model (openpyxl) and retrieve the annualized revenue growth rate ('value') as well as "Operating Scenario" bins ('bin')
	book = opx.load_workbook(valuationModel_rootAccess_directory, data_only=True)
	DCF_sheet = book['DCF Analysis']
	annualized_revenueGrowth = DCF_sheet['G32'].value
	revenueGrowth_bins = []
	for row in range(114, 119):
		rowValue = DCF_sheet.cell(row=row, column=3).value
		revenueGrowth_bins.append(rowValue)
	book.close(); book = None
	#Set the appropriate "Operating Scenario" based on which 'bin' our value is closest to, and "Revenue Growth Rate Adjustment" based on the difference between 'value' and 'bin'
	operatingScenarios = {0:'Declining', 1:'Struggling', 2:'Normal', 3:'Growth', 4:'High Growth'}
	np_revenueGrowth_bins = np.array(revenueGrowth_bins)
	bin_index = np.digitize(annualized_revenueGrowth, np_revenueGrowth_bins)
	if bin_index in (0, 5):
		if bin_index == 5:
			bin_index = 4
		operatingScenario = operatingScenarios[bin_index]
		revenueGrowth_adjustment = 0
	else:
		lowerBound_difference = annualized_revenueGrowth - revenueGrowth_bins[bin_index-1]
		upperBound_difference = revenueGrowth_bins[bin_index] - annualized_revenueGrowth
		if lowerBound_difference < upperBound_difference:
			operatingScenario = operatingScenarios[bin_index-1]
			revenueGrowth_adjustment = lowerBound_difference
		else:
			operatingScenario = operatingScenarios[bin_index]
			revenueGrowth_adjustment = -upperBound_difference
	#For revaluation only: Extract the competitive advantage and senior management gradings from our existing xlsx file (so that we can copy them to the new file later)
	if valuation_type == 'Revaluation':
		try:
			existing_file = f'{stock_directory}/{stock_ticker} Valuation.xlsx'
			book = opx.load_workbook(existing_file)
		except:			#to handle the error if file is not available (for save == 'N' revaluation stocks)
			dataExtraction_status = 'N'			#to indicate that we have not undergone data extraction for our 'subjective gradings'
		else:
			DCF_sheet = book['DCF Analysis']
			pricingPower = DCF_sheet['G3'].value
			clarity = DCF_sheet['G7'].value
			integrity = DCF_sheet['G8'].value
			leadership = DCF_sheet['G9'].value
			dataExtraction_status = 'Y'
			book.close(); book = None

	#Re-open without "data_only=True" (so that openpyxl will save WITH formulas), input the appropriate "Operating Scenario" and "Revenue Growth Rate Adjustment", then save it to our 'placeholder excel file' (so that we can open it later by xlwings)
	#NB: For save == 'Y' revaluation stocks only, also input the existing gradings into our 'placeholder excel file'
	book = opx.load_workbook(valuationModel_rootAccess_directory)
	DCF_sheet = book['DCF Analysis']
	DCF_sheet['C6'].value = operatingScenario
	DCF_sheet['C7'].value = revenueGrowth_adjustment
	if valuation_type == 'Revaluation':
		if dataExtraction_status == 'Y':
			DCF_sheet['G3'].value = pricingPower
			DCF_sheet['G7'].value = clarity
			DCF_sheet['G8'].value = integrity
			DCF_sheet['G9'].value = leadership
	book.save(valuationModel_rootAccess_directory)
	book.close(); book = None
	#'Manually' open-save-close through xlwings (so that excel can recalculate the formulas based on our changes made), re-opens by openpyxl, then extracts the stock's valuation range and "Premium (Discount)" figure of the stock
	open_excel_by_xlwings(valuationModel_rootAccess_directory)
	book = opx.load_workbook(valuationModel_rootAccess_directory, data_only=True)
	valuation_sheet = book['Valuation']
	valuation_lowRange = valuation_sheet['F12'].value
	valuation_highRange = valuation_sheet['G12'].value
	premiumDiscount = valuation_sheet['F13'].value
	book.close(); book = None

	#Set "newStatus" to either 'Shortlisted' or 'Skipped' based on the extracted "Premium (Discount)" figure, and set "shortlist_condition" and "save_condition"
	try:
		premiumDiscount_float = float(premiumDiscount)
	except:
		shortlist_condition = (premiumDiscount == 'Fair Valued')
		save_condition = shortlist_condition		#only save if the stock is 'Fair Valued' (as the other possible str entry is 'Undetermined', which needs to be skipped)
		save = 'Y' if save_condition else 'N'
	else:
		shortlist_condition = (premiumDiscount_float < 0.2)
		save_condition = (premiumDiscount_float < 0) or (shortlist_condition and existing_status == 'Confirmed')		#only save if the stock is undervalued, or is not more than 20% overvalued while having 'Confirmed' as its "existing_status" ('Shortlisted' 15% will not be saved, 'Confirmed' 25% will not be saved, 'Confirmed' 15% will be saved)
		save = 'Y' if save_condition else 'N'
	newStatus = 'Shortlisted' if shortlist_condition else 'Skipped'
	#Saving the stock valuation model if "save" is 'Y' (when shortlisted stock is undervalued or fair valued, or has been manually 'confirmed' in the past)
	#NB: If "valuation_type" is 'Revaluation', save it to existing directory. If not (which by default "valuation_type" is 'New Valuation'), save it to a new directory
	save_directory = f'{stock_directory}/{stock_ticker} Valuation.xlsx'
	if save == 'Y':
		try:
			os.makedirs(stock_directory)
		except:			#we want to handle the error if "stock_directory" already exists (for save == 'Y' revaluation stocks)
			pass
		book = opx.load_workbook(valuationModel_rootAccess_directory)
		book.save(save_directory)
		book.close(); book = None
	elif save == 'N':
		delete_directory_if_revaluation_stock(stock_directory)
	#Returns the relevant data to be appended to stock database
	relevantData = (newStatus, stock_fiscalYear, valuation_lowRange, valuation_highRange, stock_price)
	return relevantData

#Function to decide on whether a stock will be revalued on the current valuation date based on its fiscal year. Must be 3-6 months (3 < x < 6) prior to current valuation date ("x"); ex: fiscal year June-May, if August is our valuation date, and since May is 3 months prior to August, hence the stock is revalued)
def revaluationPrompt(stock_fiscalYear):
	startPos = stock_fiscalYear.find('-')
	length = len(stock_fiscalYear)
	fiscalYear = stock_fiscalYear[startPos+1:length]
	now = dt.datetime.now()
	formatted_year = now.strftime('%Y')
	formatted_fiscalYear = f'{fiscalYear} {formatted_year}'
	formatted = dt.datetime.strptime(formatted_fiscalYear, '%B %Y')
	if now < formatted:
		formatted_fiscalYear = f'{fiscalYear} {int(formatted_year)-1}'			#ex: now is Sep 2021, formatted is Dec 2021; since now < formatted, prev fiscal year is Dec 2020
		formatted = dt.datetime.strptime(formatted_fiscalYear, '%B %Y')
		prev_fiscalYear = formatted
	elif now > formatted:			#ex: now is Sep 2021, formatted is Aug 2021; since now > formatted, prev fiscal year is Aug 2021
		prev_fiscalYear = formatted
	delta = now - prev_fiscalYear
	if (delta > dt.timedelta(days=90)) and (delta < dt.timedelta(days=180)):
		return 'Y'
	else:
		return 'N'
#Function: The Main Valuation Refresh Program (updates stock database based on the results of the function "stockValuation")
def stock_valuationRefresh(index, stock_name, stock_ticker, country_code, valuation_type='New Valuation', existing_status=None):
	stockValuationRefresh_by_row = np.array(df_existingstockData.loc[index])		#copies the subject stock iteration's existing data into a numpy array
	def clearColumns_if_valuation_is_skipped():
		stockValuationRefresh_by_row[additionalNotes_index] = np.nan
		stockValuationRefresh_by_row[fiscalYear_index] = np.nan
		stockValuationRefresh_by_row[valuationLow_index] = np.nan
		stockValuationRefresh_by_row[valuationHigh_index] = np.nan
		stockValuationRefresh_by_row[previousClose_index] = np.nan
	if valuation_type == 'Price Refresh':			#same structure like our extract data function "pullData_from_WSJ"
		browser = wd.Chrome(chromedriver_filepath, desired_capabilities=capa)
		link = f'https://www.wsj.com/market-data/quotes/{country_code}/{stock_ticker}/financials/annual/income-statement'
		def forceStop_handler():
			wait = WebDriverWait(browser, 5)
			browser.get(link)
			try:
				wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="navCol"]/div[2]')))
			except TimeoutException:
				pass
			else:
				time.sleep(5)
			browser.execute_script("window.stop();")
		for attempts in range(8):			#we allow up to 8 attempts to fetch the stock price (as well as fiscal year data)
			try:
				stockPrice_raw = browser.find_element_by_xpath('//*[@id="quote_val"]').text
				fiscalYear_raw = browser.find_element_by_xpath('//*[@id="cr_cashflow"]/div[2]/div/table/thead/tr/th[1]').text
			except (NoSuchElementException, StaleElementReferenceException):
				try:
					browser.find_element_by_xpath('//*[@id="main"]/div/div/div/div[1]')
				except (NoSuchElementException, StaleElementReferenceException):
					forceStop_handler()
				else:
					browser.quit()
					raise CannotFetchDataException
			else:
				stockPrice = float(stockPrice_raw.replace(',', ''))
				if country_code == 'UK':
					stockPrice = stockPrice / 100 			#specially for UK, as their stock prices are denominated in pence instead of GBP
				fiscalYear_regex = re.compile(r"\w+[-]\w+")
				fiscalYear_result = fiscalYear_regex.findall(fiscalYear_raw)
				fiscalYear = fiscalYear_result[0]
				stockValuationRefresh_by_row[additionalNotes_index] = np.nan
				stockValuationRefresh_by_row[fiscalYear_index] = fiscalYear
				stockValuationRefresh_by_row[previousClose_index] = stockPrice
				browser.quit()
				break
		else:
			stockValuationRefresh_by_row[additionalNotes_index] = priceRefresh_CannotFetchDataException_message
			browser.quit()
	else:
		try:
			newStatus, fiscalYear, valuation_lowRange, valuation_highRange, stock_price = stockValuation(stock_name, stock_ticker, country_code, valuation_type=valuation_type, existing_status=existing_status)
		except CannotFetchDataException:
			if valuation_type == 'Revaluation':
				stockValuationRefresh_by_row[additionalNotes_index] = revaluation_CannotFetchDataException_message
			elif valuation_type == 'New Valuation':
				stockValuationRefresh_by_row[additionalNotes_index] = newValuation_CannotFetchDataException_message
		except SkipValuationException:
			newStatus = 'Skipped'
			stockValuationRefresh_by_row[status_index] = newStatus
			clearColumns_if_valuation_is_skipped()
		else:
			if newStatus == 'Shortlisted':
				if valuation_type == 'New Valuation':		#we only change 'Status' column if its a new valuation that returns "Shortlisted" as its status; if revaluation returns "Shortlisted", follow the previous predecent (either "Confirmed" or also "Shortlisted")
					stockValuationRefresh_by_row[status_index] = newStatus
				stockValuationRefresh_by_row[additionalNotes_index] = np.nan
				stockValuationRefresh_by_row[fiscalYear_index] = fiscalYear
				stockValuationRefresh_by_row[valuationLow_index] = valuation_lowRange
				stockValuationRefresh_by_row[valuationHigh_index] = valuation_highRange
				stockValuationRefresh_by_row[previousClose_index] = stock_price
			elif newStatus == 'Skipped':
				stockValuationRefresh_by_row[status_index] = newStatus
				clearColumns_if_valuation_is_skipped()
	return stockValuationRefresh_by_row

#Specifying directories, the webdriver's filepath, and setting up browser ("capa" to enable the option to force stop loading page) 
directory = '/Users/delvinkennedy/sandbox/Personal Investments/Stocks'
stockDatabase_directory = f'{directory}/Stock Database & Valuation Models/Stock Database.xlsx'
valuationTemplate_directory = f'{directory}/Stock Database & Valuation Models/Stock Valuation.xlsx'
valuationModel_rootAccess_directory = f'{directory}/Stock Database & Valuation Models/Stock Valuation (Root).xlsx'
chromedriver_filepath = '/Users/delvinkennedy/sandbox/Python/chromedriver'
capa = DesiredCapabilities.CHROME; capa["pageLoadStrategy"] = "none"		#this is default syntax to enable 'force stop' functionality
#Loading stock database, and setting up global variables for the loop
priceRefresh_CannotFetchDataException_message = 'Retrieve stock price pending: Cannot fetch data from WSJ.'
revaluation_CannotFetchDataException_message = 'Revaluation pending: Cannot fetch data from WSJ.'
newValuation_CannotFetchDataException_message = 'New valuation pending: Cannot fetch data from WSJ.'
countryType_startingRowIndex = 8; countryType_startingColIndex = 1; countryType_endingColIndex = 2
countryType_table = pd.read_excel(stockDatabase_directory, skiprows=range(0, countryType_startingRowIndex-1), usecols=range(countryType_startingColIndex, countryType_endingColIndex+1))
stockDatabase_startingRowIndex = 32; stockDatabase_startingColIndex = 1; stockDatabase_endingColIndex = 9
df_existingstockData = pd.read_excel(stockDatabase_directory, skiprows=range(0, stockDatabase_startingRowIndex-1), usecols=range(stockDatabase_startingColIndex, stockDatabase_endingColIndex+1))
df_existingstockData = df_existingstockData.dropna(axis='index', how='all', subset=['Stock Name', 'Stock Ticker', 'Country'])
status_index = df_existingstockData.columns.get_loc('Status')
additionalNotes_index = df_existingstockData.columns.get_loc('Additional Notes')
fiscalYear_index = df_existingstockData.columns.get_loc('Fiscal Year')
valuationLow_index = df_existingstockData.columns.get_loc('Valuation (Low)')
valuationHigh_index = df_existingstockData.columns.get_loc('Valuation (High)')
previousClose_index = df_existingstockData.columns.get_loc('Previous Close')

#Prompts for choice (normal vs clean-up): Normal will iterate over 'Status' column, while clean-up will iterate over 'Additional Notes' column
choice = pyip.inputYesNo(prompt='Select type of valuation:\n1) Normal (please input N into the console)\n2) Clean-up (please input C into the console)\nYour choice: ', yesVal='N', noVal='C')
#Loops stock database to automate stock valuation based on inputted choice, and append new updated stock database into a list
def mouse_movement(index):
	if index % 2 == 0:
		pyag.move(1, 0, duration=0.25)			#if index is even number, relative movement by +1 pixel (to the right)
	else:
		pyag.move(-1, 0, duration=0.25)			#if index is odd number, relative movement by -1 pixel (to the left)
num_of_rows, num_of_cols = df_existingstockData.shape
stockDatabase = []
for index in tqdm(range(num_of_rows), desc='Progress'):
	stock_name = df_existingstockData.loc[index, 'Stock Name']
	stock_ticker = df_existingstockData.loc[index, 'Stock Ticker']
	country_code = df_existingstockData.loc[index, 'Country']
	status = df_existingstockData.loc[index, 'Status']
	if choice == 'N':
		if status in ('Confirmed', 'Shortlisted'):
			mouse_movement(index)			#so that our laptop doesn't go to sleep, as our program won't run if it does (even screensaver mode is also a no-go); we don't need to call this for "Skipped" stocks, as they won't consume any time at all to iterate (putting it there will just needlessly slow our program)
			stock_fiscalYear = df_existingstockData.loc[index, 'Fiscal Year']
			revaluation_prompt = revaluationPrompt(stock_fiscalYear)
			if revaluation_prompt == 'Y':
				stockData_by_row = stock_valuationRefresh(index, stock_name, stock_ticker, country_code, valuation_type='Revaluation', existing_status=status)			#specially for 'Revaluation', we need to 'bring in' "status" to our function (since it will be used to decide on "save_condition" later on)
			elif revaluation_prompt == 'N':
				stockData_by_row = stock_valuationRefresh(index, stock_name, stock_ticker, country_code, valuation_type='Price Refresh')
		elif status != 'Skipped':
			mouse_movement(index)
			stockData_by_row = stock_valuationRefresh(index, stock_name, stock_ticker, country_code)
		elif status == 'Skipped':
			stockData_by_row = np.array(df_existingstockData.loc[index])
	elif choice == 'C':
		additionalNotes = df_existingstockData.loc[index, 'Additional Notes']
		if additionalNotes == priceRefresh_CannotFetchDataException_message:
			mouse_movement(index)
			stockData_by_row = stock_valuationRefresh(index, stock_name, stock_ticker, country_code, valuation_type='Price Refresh')
		elif additionalNotes == revaluation_CannotFetchDataException_message:
			mouse_movement(index)
			stockData_by_row = stock_valuationRefresh(index, stock_name, stock_ticker, country_code, valuation_type='Revaluation', existing_status=status)
		elif additionalNotes == newValuation_CannotFetchDataException_message:
			mouse_movement(index)
			stockData_by_row = stock_valuationRefresh(index, stock_name, stock_ticker, country_code)
		else:
			stockData_by_row = np.array(df_existingstockData.loc[index])
	stockDatabase.append(stockData_by_row)
	print(stockData_by_row)			#just for 'tracking' purposes (so we can 'track' the progress per stock and see whether our program is still running as expected)
#Organizes the list "stockDatabase" into a single dataframe
numpy_stockDatabase = np.vstack(stockDatabase)
df_stockDatabase = pd.DataFrame(numpy_stockDatabase)
#Gets the current date in excel date format
def format_datetime_to_excel(secs_since_pythonEpoch):
	excelDay_at_pythonEpoch = 25569
	pythonDay_since_pythonEpoch = round(secs_since_pythonEpoch / 86400, 0)
	excelDay = excelDay_at_pythonEpoch + pythonDay_since_pythonEpoch
	return excelDay
now = time.time()
currentDate_excelFormat = format_datetime_to_excel(now)
#Appends the current valuation date as well as "df_stockDatabase" to "Stock Database.xlsx"
book = opx.load_workbook(stockDatabase_directory)
sheet = book['Stock Database']
sheet['C2'].value = currentDate_excelFormat
writer = pd.ExcelWriter(stockDatabase_directory, engine='openpyxl')
writer.book = book; writer.sheets = {ws.title: ws for ws in book.worksheets}
for sheetname in writer.sheets:
	df_stockDatabase.to_excel(writer, sheet_name='Stock Database', startrow=stockDatabase_startingRowIndex, startcol=stockDatabase_startingColIndex, index=False, header=False)
writer.save()
writer = None; book = None
print('Program ran successfully. Stock database has been updated. Files and folders have been updated.')
